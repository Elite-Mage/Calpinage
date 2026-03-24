#!/usr/bin/env python3
"""
parse_dxf.py — Convertit un fichier DXF (ou DWG via dwg2dxf) en JSON + Excel
              structurés pour l'appli Calepinage.

Usage :
  python3 parse_dxf.py fichier.dxf                       → JSON sur stdout
  python3 parse_dxf.py fichier.dxf --excel sortie.xlsx   → JSON stdout + Excel
  python3 parse_dxf.py fichier.dxf --out sortie.json     → JSON fichier + Excel auto
  python3 parse_dxf.py fichier.dwg --excel sortie.xlsx   → conversion DWG→DXF puis pareil

Dépendances : ezdxf, openpyxl  (pip install ezdxf openpyxl)
Pour DWG   : LibreDWG (dwg2dxf) doit être installé (apt install libredwg-tools)
"""

import sys
import json
import math
import os
import subprocess
import tempfile
import argparse
import datetime
from collections import Counter, defaultdict

try:
    import ezdxf
except ImportError:
    print("ERREUR: ezdxf manquant. Lancez: pip install ezdxf", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── Palette ACI → informations groupe ───────────────────────────────────────
# Clé = numéro de couleur ACI DXF
ACI_COLOR_MAP = {
    30:  {"name": "Panneaux Orange",    "color": "#fd9a51"},
    25:  {"name": "Panneaux Marron",    "color": "#8B5E3C"},
    1:   {"name": "Panneaux Rouge",     "color": "#e63946"},
    2:   {"name": "Panneaux Jaune",     "color": "#f4d03f"},
    3:   {"name": "Panneaux Vert",      "color": "#2ecc71"},
    4:   {"name": "Panneaux Cyan",      "color": "#1abc9c"},
    5:   {"name": "Panneaux Bleu",      "color": "#3498db"},
    6:   {"name": "Panneaux Magenta",   "color": "#9b59b6"},
    7:   {"name": "Panneaux Blanc/Noir","color": "#aaaaaa"},
    114: {"name": "Panneaux Blanc/Noir","color": "#aaaaaa"},
}

def round_mm(v):
    """Arrondit à l'entier millimètre le plus proche."""
    return int(round(float(v)))

def classify_subtype_by_position(panel, all_panels_facade):
    """
    Classifie un panneau en sous-type selon sa position dans la façade.
    - Bandeau Toiture : dernier panneau en haut, rien au-dessus, h < 1600mm
    - Plein           : h ≥ 1600mm
    - Pièce spéciale  : h < 1600mm avec un panneau au-dessus
    """
    ph = round_mm(panel["ymax"] - panel["ymin"])
    if ph >= 1600:
        return "Plein"
    # Vérifier s'il y a un panneau au-dessus (chevauchement horizontal > 5mm)
    has_above = False
    for other in all_panels_facade:
        if other is panel:
            continue
        if other["ymin"] > panel["ymax"] + 20:
            overlap_x = min(panel["xmax"], other["xmax"]) - max(panel["xmin"], other["xmin"])
            if overlap_x > 5:
                has_above = True
                break
    if not has_above:
        return "Bandeau Toiture"
    return "Pièce spéciale"

def _dwg_to_dxf_libredwg(dwg_path):
    """Conversion DWG→DXF via dwg2dxf (LibreDWG) si disponible."""
    for cmd in ["dwg2dxf", "/usr/bin/dwg2dxf", "/usr/local/bin/dwg2dxf"]:
        result = subprocess.run(["which", cmd.split("/")[-1]], capture_output=True, text=True)
        if result.returncode == 0:
            tmpdir = tempfile.mkdtemp()
            base = os.path.splitext(os.path.basename(dwg_path))[0]
            out_dxf = os.path.join(tmpdir, base + ".dxf")
            subprocess.run([cmd, dwg_path, "-o", out_dxf], check=True,
                           capture_output=True, text=True)
            return out_dxf
    return None


def _dwg_to_dxf_aspose(dwg_path):
    """Conversion DWG→DXF via aspose-cad (fallback)."""
    try:
        import aspose.cad as cad
        from aspose.cad.imageoptions import DxfOptions
        image = cad.Image.load(dwg_path)
        opts = DxfOptions()
        tmpdir = tempfile.mkdtemp()
        base = os.path.splitext(os.path.basename(dwg_path))[0]
        out_dxf = os.path.join(tmpdir, base + ".dxf")
        image.save(out_dxf, opts)
        return out_dxf
    except ImportError:
        return None
    except Exception as e:
        print(f"aspose-cad conversion failed: {e}", file=sys.stderr)
        return None


def dwg_to_dxf(dwg_path):
    """
    Convertit DWG→DXF en essayant d'abord dwg2dxf (LibreDWG),
    puis aspose-cad en fallback.
    Retourne le chemin du DXF temporaire ou lève une exception.
    """
    # Essai 1 : LibreDWG (dwg2dxf)
    result = _dwg_to_dxf_libredwg(dwg_path)
    if result:
        return result

    # Essai 2 : aspose-cad
    result = _dwg_to_dxf_aspose(dwg_path)
    if result:
        return result

    raise RuntimeError(
        "Aucun convertisseur DWG→DXF disponible.\n"
        "Installez l'un des deux :\n"
        "  pip install aspose-cad\n"
        "  ou compilez LibreDWG depuis https://github.com/LibreDWG/libredwg"
    )

# ─── Calcul ossature par façade (analyse spatiale) ────────────────────────────

def calc_ossature_facades(rects_spatial, facade_labels, entraxe_max=600, panel_facade_map=None):
    """
    Analyse la disposition spatiale des panneaux pour calculer l'ossature
    (Oméga et Zed) par façade.

    Logique :
    - OMÉGA = jonction entre 2 panneaux adjacents (gap ≤ 20mm)
      - Toujours continu (traverse les joints inter-étages)
      - Rectangle 80mm de large
    - ZED = bord libre (ouverture) + entraxe support
      - Pas de ZED aux bords de façade (gauche/droite)
      - Entraxe régulier : 600mm, bandeau toiture : 800mm
      - Bord libre : aux ouvertures (fenêtres, portes)
      - Rectangle 40mm de large

    Retourne un dict {facade_name: {"omega_mm", "zed_mm", "omega_ml", "zed_ml",
                                     "omega_details": {h: qty}, "zed_details": {h: qty}}}
    """
    JOINT = 8
    MAX_GAP = 20
    BANDEAU_TOITURE_MAX_H = 1600
    ENTRAXE_BANDEAU = 800

    def nearest_facade_oss(xcenter, ycenter):
        if isinstance(facade_labels[0][0], (tuple, list)):
            return min(facade_labels, key=lambda lbl: (lbl[0][0] - xcenter)**2 + (lbl[0][1] - ycenter)**2)[1]
        return min(facade_labels, key=lambda lbl: abs(lbl[0] - xcenter))[1]

    def get_gaps(left_intervals, right_intervals):
        """Get Y segments where one side has panel but other doesn't."""
        events = []
        for s, e in left_intervals:
            events.append((s, "L+"))
            events.append((e, "L-"))
        for s, e in right_intervals:
            events.append((s, "R+"))
            events.append((e, "R-"))
        events.sort(key=lambda x: (x[0], x[1]))
        left_gaps, right_gaps = [], []
        l_count = r_count = 0
        prev_y = None
        for y, ev in events:
            if prev_y is not None and y > prev_y:
                if l_count > 0 and r_count == 0:
                    right_gaps.append((prev_y, y))
                elif r_count > 0 and l_count == 0:
                    left_gaps.append((prev_y, y))
            if ev == "L+":
                l_count += 1
            elif ev == "L-":
                l_count -= 1
            elif ev == "R+":
                r_count += 1
            elif ev == "R-":
                r_count -= 1
            prev_y = y
        return left_gaps, right_gaps

    by_facade = defaultdict(list)
    for i, r in enumerate(rects_spatial):
        if panel_facade_map and i in panel_facade_map:
            fname = panel_facade_map[i]
        else:
            cx = (r["xmin"] + r["xmax"]) / 2
            cy = (r["ymin"] + r["ymax"]) / 2
            fname = nearest_facade_oss(cx, cy)
        by_facade[fname].append(r)

    result = {}
    for fname in sorted(by_facade.keys()):
        all_panels = by_facade[fname]

        # Classify bandeau toiture (top, nothing above, h < 1600mm) for entraxe only
        bandeau_set = set()
        for p in all_panels:
            ph = round(p["ymax"] - p["ymin"])
            if ph >= BANDEAU_TOITURE_MAX_H:
                continue
            has_above = False
            for other in all_panels:
                if other is p:
                    continue
                if other["ymin"] > p["ymax"] + 20:
                    overlap_x = min(p["xmax"], other["xmax"]) - max(p["xmin"], other["xmin"])
                    if overlap_x > 5:
                        has_above = True
                        break
            if not has_above:
                bandeau_set.add(id(p))

        total_omega_mm = 0
        total_zed_mm = 0
        omega_details = defaultdict(int)
        zed_details = defaultdict(int)

        def add_omega(h):
            nonlocal total_omega_mm
            total_omega_mm += h
            omega_details[h] += 1

        def add_zed(h):
            nonlocal total_zed_mm
            total_zed_mm += h
            zed_details[h] += 1

        # --- Entraxe ZED (per panel, based on width) ---
        for p in all_panels:
            pw = round(p["xmax"] - p["xmin"])
            ph = round(p["ymax"] - p["ymin"])
            ent = ENTRAXE_BANDEAU if id(p) in bandeau_set else entraxe_max
            nb_zed = max(0, math.ceil(pw / ent) - 1)
            for _ in range(nb_zed):
                add_zed(ph)

        # --- Pairwise junction detection for omega and gap-zed ---
        # Find all right-edge → left-edge adjacencies
        junctions = defaultdict(lambda: {"left": [], "right": []})
        has_right_neighbor = set()
        has_left_neighbor = set()
        for p in all_panels:
            for q in all_panels:
                if q is p:
                    continue
                gap = q["xmin"] - p["xmax"]
                if -5 <= gap <= MAX_GAP and q["xmin"] > p["xmin"] + 20:
                    jx = round((p["xmax"] + q["xmin"]) / 2)
                    junctions[jx]["left"].append(p)
                    junctions[jx]["right"].append(q)
                    has_right_neighbor.add(id(p))
                    has_left_neighbor.add(id(q))

        # Process each junction: omega where both sides overlap, zed in gaps
        for jx in sorted(junctions.keys()):
            left_panels = junctions[jx]["left"]
            right_panels = junctions[jx]["right"]

            # Omega: Y-overlaps between left and right panels, merge across joints
            overlaps = []
            for pl in left_panels:
                for pr in right_panels:
                    ov_start = max(pl["ymin"], pr["ymin"])
                    ov_end = min(pl["ymax"], pr["ymax"])
                    if ov_end - ov_start > 1:
                        overlaps.append((ov_start, ov_end))
            overlaps.sort()
            if overlaps:
                merged_ov = [list(overlaps[0])]
                for ov in overlaps[1:]:
                    prev = merged_ov[-1]
                    if ov[0] <= prev[1] + MAX_GAP:
                        prev[1] = max(prev[1], ov[1])
                    else:
                        merged_ov.append(list(ov))
                for start, end in merged_ov:
                    add_omega(round(end - start))

            # Gap ZED: Y-ranges where one side has panel but other doesn't
            left_intervals = list({(round(p["ymin"]), round(p["ymax"])) for p in left_panels})
            right_intervals = list({(round(p["ymin"]), round(p["ymax"])) for p in right_panels})
            left_gaps, right_gaps = get_gaps(left_intervals, right_intervals)
            for gs, ge in right_gaps:
                h = round(ge - gs)
                if h >= 100:
                    add_zed(h)
            for gs, ge in left_gaps:
                h = round(ge - gs)
                if h >= 100:
                    add_zed(h)

        # --- Free edge ZED: interior panel edges with no neighbor ---
        for p in all_panels:
            if id(p) not in has_right_neighbor:
                # No right neighbor — is this an interior free edge (opening)?
                if any(q["xmin"] > p["xmax"] + MAX_GAP for q in all_panels if q is not p):
                    add_zed(round(p["ymax"] - p["ymin"]))
            if id(p) not in has_left_neighbor:
                if any(q["xmax"] < p["xmin"] - MAX_GAP for q in all_panels if q is not p):
                    add_zed(round(p["ymax"] - p["ymin"]))

        result[fname] = {
            "omega_mm": round(total_omega_mm),
            "zed_mm": round(total_zed_mm),
            "omega_ml": round(total_omega_mm / 1000, 2),
            "zed_ml": round(total_zed_mm / 1000, 2),
            "omega_details": dict(omega_details),
            "zed_details": dict(zed_details),
        }

    return result


def parse_dxf_file(filepath):
    """
    Lit un fichier DXF et retourne les données structurées sous forme de dict
    compatible avec le format JSON de l'appli Calepinage.
    """
    doc = ezdxf.readfile(filepath)
    msp = doc.modelspace()

    # 1. Récupère les labels de façades (entités TEXT)
    facade_labels = []
    for e in msp:
        if e.dxftype() in ("TEXT", "MTEXT"):
            try:
                txt = e.dxf.text if e.dxftype() == "TEXT" else e.text
                x = e.dxf.insert[0]
                facade_labels.append((x, txt.strip()))
            except Exception:
                pass

    facade_labels.sort(key=lambda t: t[0])

    # Si pas de texte, on numérotera les façades après extraction des rectangles
    auto_number_facades = len(facade_labels) == 0

    # 2. Construit un dict layer → couleur ACI pour résoudre BYLAYER
    layer_colors = {}
    for layer in doc.layers:
        try:
            c = layer.dxf.color
            if c is not None:
                layer_colors[layer.dxf.name] = abs(c)
        except Exception:
            pass

    def resolve_color(entity):
        """Résout la couleur ACI : entité > calque (BYLAYER)."""
        color = entity.dxf.get("color", 256)
        if color == 256:  # BYLAYER
            layer_name = entity.dxf.get("layer", "0")
            color = layer_colors.get(layer_name, 256)
        return color

    # 3. Extrait tous les rectangles (LWPOLYLINE + POLYLINE)
    rects = []
    rects_spatial = []  # Garde les positions complètes pour l'analyse ossature

    def add_rect(xmin, xmax, ymin, ymax, color_aci):
        """Ajoute un rectangle avec dédoublonnage spatial (tolérance 5mm)."""
        # Remap ACI 114 → 7 (blanc)
        if color_aci == 114:
            color_aci = 7
        raw_w = xmax - xmin
        raw_h = ymax - ymin
        if raw_w < 10 or raw_h < 10:
            return
        # Dédoublonnage spatial
        TOL = 5
        for ex in rects_spatial:
            if (abs(ex["xmin"] - xmin) < TOL and abs(ex["xmax"] - xmax) < TOL and
                    abs(ex["ymin"] - ymin) < TOL and abs(ex["ymax"] - ymax) < TOL):
                return  # doublon
        w = round_mm(raw_w)  # L = horizontal
        h = round_mm(raw_h)  # H = vertical
        xcenter = (xmin + xmax) / 2.0
        rects.append({"xcenter": xcenter, "w": w, "h": h, "color": color_aci})
        rects_spatial.append({"xmin": xmin, "xmax": xmax, "ymin": ymin, "ymax": ymax, "color": color_aci})

    all_lines = []  # Collect LINEs from modelspace AND blocks

    def process_entity(e, offset_x=0.0, offset_y=0.0, parent_color=None, parent_layer=None):
        """Traite une entité DXF (supporte offset pour les INSERT/blocks)."""
        def _resolve(e):
            c = resolve_color(e)
            if c == 0:  # BYBLOCK
                if parent_color:
                    c = parent_color
                else:
                    # BYBLOCK in modelspace (no parent INSERT) → resolve to layer color
                    layer_name = e.dxf.get("layer", "0")
                    c = layer_colors.get(layer_name, 7)
            if c == 256 and parent_layer:
                c = layer_colors.get(parent_layer, 256)
            return c

        if e.dxftype() == "LWPOLYLINE":
            pts = [(p[0] + offset_x, p[1] + offset_y) for p in e.get_points()]
            if len(pts) >= 3:
                xs = [float(p[0]) for p in pts]
                ys = [float(p[1]) for p in pts]
                add_rect(min(xs), max(xs), min(ys), max(ys), _resolve(e))
        elif e.dxftype() == "POLYLINE":
            try:
                pts = [(v.dxf.location.x + offset_x, v.dxf.location.y + offset_y) for v in e.vertices]
                if len(pts) >= 3:
                    xs = [float(p[0]) for p in pts]
                    ys = [float(p[1]) for p in pts]
                    add_rect(min(xs), max(xs), min(ys), max(ys), _resolve(e))
            except Exception:
                pass
        elif e.dxftype() == "LINE":
            color_aci = _resolve(e)
            layer = e.dxf.get("layer", "0")
            x1 = float(e.dxf.start.x) + offset_x
            y1 = float(e.dxf.start.y) + offset_y
            x2 = float(e.dxf.end.x) + offset_x
            y2 = float(e.dxf.end.y) + offset_y
            all_lines.append((x1, y1, x2, y2, color_aci, layer))
        elif e.dxftype() == "INSERT":
            try:
                block_name = e.dxf.name
                block = doc.blocks.get(block_name)
                if block is None:
                    return
                ins_x = float(e.dxf.get("insert", (0, 0, 0))[0]) + offset_x
                ins_y = float(e.dxf.get("insert", (0, 0, 0))[1]) + offset_y
                ins_color = resolve_color(e)
                ins_layer = e.dxf.get("layer", "0")
                for be in block:
                    process_entity(be, ins_x, ins_y, ins_color, ins_layer)
            except Exception:
                pass

    for e in msp:
        process_entity(e)

    # 2b. Assembler les LINE isolées en rectangles (4 lignes → 1 rectangle)
    lines_by_key = defaultdict(list)
    for line in all_lines:
        x1, y1, x2, y2, color_aci, layer = line
        key = f"{color_aci}|{layer}"
        lines_by_key[key].append((x1, y1, x2, y2, color_aci, layer))

    for key, segs in lines_by_key.items():
        h_segs = [(s[0], s[1], s[2], s[3]) for s in segs if abs(s[1] - s[3]) < 1]
        v_segs = [(s[0], s[1], s[2], s[3]) for s in segs if abs(s[0] - s[2]) < 1]
        color_aci = segs[0][4]
        for hi in range(len(h_segs)):
            h1 = h_segs[hi]
            h1xmin, h1xmax = min(h1[0], h1[2]), max(h1[0], h1[2])
            for hj in range(hi + 1, len(h_segs)):
                h2 = h_segs[hj]
                h2xmin, h2xmax = min(h2[0], h2[2]), max(h2[0], h2[2])
                if abs(h1xmin - h2xmin) > 2 or abs(h1xmax - h2xmax) > 2:
                    continue
                ymin = min(h1[1], h2[1])
                ymax = max(h1[1], h2[1])
                if ymax - ymin < 10:
                    continue
                has_left = has_right = False
                for v in v_segs:
                    vx = (v[0] + v[2]) / 2
                    vymin, vymax = min(v[1], v[3]), max(v[1], v[3])
                    if abs(vymin - ymin) < 2 and abs(vymax - ymax) < 2:
                        if abs(vx - h1xmin) < 2:
                            has_left = True
                        if abs(vx - h1xmax) < 2:
                            has_right = True
                if has_left and has_right:
                    add_rect(h1xmin, h1xmax, ymin, ymax, color_aci)

    # 2c. Dédoublonnage englobé : supprimer les petits panneaux contenus dans un plus grand de même couleur
    englobed = set()
    for i, r in enumerate(rects_spatial):
        if i in englobed:
            continue
        r_area = (r["xmax"] - r["xmin"]) * (r["ymax"] - r["ymin"])
        for j, o in enumerate(rects_spatial):
            if i == j or j in englobed:
                continue
            if o["color"] != r["color"]:
                continue
            o_area = (o["xmax"] - o["xmin"]) * (o["ymax"] - o["ymin"])
            if o_area >= r_area:
                continue
            if (o["xmin"] >= r["xmin"] - 2 and o["xmax"] <= r["xmax"] + 2 and
                    o["ymin"] >= r["ymin"] - 2 and o["ymax"] <= r["ymax"] + 2):
                englobed.add(j)
    if englobed:
        rects[:] = [r for i, r in enumerate(rects) if i not in englobed]
        rects_spatial[:] = [r for i, r in enumerate(rects_spatial) if i not in englobed]

    # 3. Auto-numérotation des façades si pas de texte dans le DXF
    #    Assigne directement chaque panneau à sa façade via Y-row + X-gap
    panel_facade_map = {}  # index rects_spatial → nom façade
    if auto_number_facades:
        if rects_spatial:
            # Étape 1 : grouper en rangées Y (panneaux connectés avec tolérance 20mm)
            indexed = list(enumerate(rects_spatial))
            indexed.sort(key=lambda x: x[1]["ymin"])
            y_rows = []
            cur_row = [indexed[0]]
            cur_ymax = indexed[0][1]["ymax"]
            for item in indexed[1:]:
                if item[1]["ymin"] <= cur_ymax + 20:
                    cur_ymax = max(cur_ymax, item[1]["ymax"])
                    cur_row.append(item)
                else:
                    y_rows.append(cur_row)
                    cur_row = [item]
                    cur_ymax = item[1]["ymax"]
            y_rows.append(cur_row)

            # Étape 2 : dans chaque rangée Y, détecter les gaps X réels (sweep line)
            facade_labels = []
            facade_idx = 1
            for row_items in y_rows:
                row_items.sort(key=lambda x: x[1]["xmin"])
                sub_groups = [[row_items[0]]]
                max_xmax = row_items[0][1]["xmax"]
                for item in row_items[1:]:
                    if item[1]["xmin"] > max_xmax + 50:
                        sub_groups.append([item])
                    else:
                        sub_groups[-1].append(item)
                    max_xmax = max(max_xmax, item[1]["xmax"])

                for sg in sub_groups:
                    fname = f"Façade {facade_idx}"
                    cx = (min(r["xmin"] for _, r in sg) + max(r["xmax"] for _, r in sg)) / 2
                    cy = (min(r["ymin"] for _, r in sg) + max(r["ymax"] for _, r in sg)) / 2
                    facade_labels.append(((cx, cy), fname))
                    for idx, _ in sg:
                        panel_facade_map[idx] = fname
                    facade_idx += 1
        else:
            facade_labels = [((0.0, 0.0), "Façade 1")]

    # nearest_facade pour les labels TEXT du DXF (1D) ou fallback
    def nearest_facade(xcenter, ycenter=None):
        if not facade_labels:
            return "Façade 1"
        if isinstance(facade_labels[0][0], (tuple, list)):
            if ycenter is not None:
                return min(facade_labels, key=lambda lbl: (lbl[0][0] - xcenter)**2 + (lbl[0][1] - ycenter)**2)[1]
            return min(facade_labels, key=lambda lbl: abs(lbl[0][0] - xcenter))[1]
        return min(facade_labels, key=lambda lbl: abs(lbl[0] - xcenter))[1]

    # Fonction d'assignation : directe si auto-numéroté, sinon par proximité TEXT
    def get_facade(idx, r):
        if idx in panel_facade_map:
            return panel_facade_map[idx]
        return nearest_facade((r["xmin"] + r["xmax"]) / 2, (r["ymin"] + r["ymax"]) / 2)

    # 4. Classifie les panneaux par façade pour déterminer les sous-types par position
    panels_by_facade = defaultdict(list)
    for i, r in enumerate(rects_spatial):
        panels_by_facade[get_facade(i, r)].append(r)

    # 5. Regroupe par (couleur ACI, façade) → compteur (w, h) avec sous-type positionnel
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(Counter)))
    for i, r in enumerate(rects):
        spatial = rects_spatial[i]
        facade = get_facade(i, spatial)
        subtype = classify_subtype_by_position(spatial, panels_by_facade[facade])
        data[r["color"]][facade][subtype][(r["w"], r["h"])] += 1

    # 6. Construit la structure JSON de l'appli
    groups = []
    gid = 1
    for color_aci in sorted(data.keys()):
        info = ACI_COLOR_MAP.get(color_aci, {
            "name": f"Panneaux (ACI {color_aci})",
            "color": "#888888",
        })

        subsections = []
        ssid = 1
        for facade_name in sorted(data[color_aci].keys()):
            panel_subtypes = []
            stid = 1
            for st_name in ["Bandeau Toiture", "Plein", "Pièce spéciale"]:
                dim_counter = data[color_aci][facade_name].get(st_name, {})
                if not dim_counter:
                    continue
                pieces = [
                    {"id": pid, "w": w, "h": h, "qty": qty}
                    for pid, ((w, h), qty) in enumerate(
                        sorted(dim_counter.items(), key=lambda x: x[0][0] * x[0][1], reverse=True), 1
                    )
                ]
                panel_subtypes.append({
                    "id": stid,
                    "name": st_name,
                    "pieces": pieces,
                    "nextPieceId": len(pieces) + 1,
                })
                stid += 1

            if panel_subtypes:
                subsections.append({
                    "id": ssid,
                    "name": facade_name,
                    "panelSubtypes": panel_subtypes,
                    "nextSubtypeId": stid,
                    "activeSubtypeId": 1,
                })
                ssid += 1

        if subsections:
            groups.append({
                "id": gid,
                "name": info["name"],
                "color": info["color"],
                "reference": "",
                "coloris": "",
                "epaisseur": 18,
                "prixM2": "",
                "subsections": subsections,
                "nextSubsectionId": ssid,
                "activeSubsectionId": 1,
            })
            gid += 1

    today = datetime.date.today().isoformat()
    base_name = os.path.splitext(os.path.basename(filepath))[0]

    # Calcul ossature par façade (analyse spatiale)
    ossature = calc_ossature_facades(rects_spatial, facade_labels, panel_facade_map=panel_facade_map)

    return {
        "version": "6.0",
        "chantier": {"nom": base_name, "date": today},
        "stockPanels": [],
        "nextStockId": 1,
        "groups": groups,
        "nextGroupId": gid,
        "activeGroupId": 1,
        "ossature_facades": ossature,
        "rectsSpatial": rects_spatial,
        "facadeLabels": facade_labels,
        "panelFacadeMap": panel_facade_map,
    }

# ─── Génération Excel ─────────────────────────────────────────────────────────

def hex_to_argb(hex_color):
    """'#fd9a51' → 'FFFD9A51'"""
    h = hex_color.lstrip("#")
    return "FF" + h.upper()

def _border():
    thin = Side(style="thin", color="FFD0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_to_argb(hex_color))

def generate_excel(data, out_path):
    """Génère un fichier Excel détaillé à partir du dict data (format JSON appli)."""
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl manquant. Lancez: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # retire la feuille par défaut

    # ── 1. Feuille Récapitulatif ────────────────────────────────────────────
    ws_recap = wb.create_sheet("Récapitulatif")
    title = data["chantier"]["nom"] + " — Récapitulatif"
    ws_recap.append([title])
    ws_recap["A1"].font = Font(bold=True, size=13)

    facades_order = []
    for g in data["groups"]:
        for ss in g["subsections"]:
            if ss["name"] not in facades_order:
                facades_order.append(ss["name"])

    ws_recap.append(["Couleur", "Façade", "Sous-type", "Larg. (mm)", "Haut. (mm)", "Quantité", "Surface unit. (m²)", "Surface totale (m²)"])
    for col in range(1, 9):
        ws_recap.cell(2, col).font = Font(bold=True)
        ws_recap.cell(2, col).fill = PatternFill("solid", fgColor="FFE0E0E0")
        ws_recap.cell(2, col).border = _border()

    grand_total = 0.0
    for g in data["groups"]:
        fill = _header_fill(g["color"])
        for ss in g["subsections"]:
            for st in ss["panelSubtypes"]:
                for p in st["pieces"]:
                    c1, c2 = p["w"], p["h"]
                    s_unit = round(c1 * c2 / 1_000_000, 4)
                    s_tot = round(s_unit * p["qty"], 4)
                    grand_total += s_tot
                    ws_recap.append([g["name"], ss["name"], st["name"], c1, c2, p["qty"], s_unit, s_tot])
                    r = ws_recap.max_row
                    for col in range(1, 9):
                        ws_recap.cell(r, col).border = _border()
                    ws_recap.cell(r, 1).fill = fill

    ws_recap.append(["TOTAL", None, None, None, None, None, None, round(grand_total, 4)])
    r = ws_recap.max_row
    for col in range(1, 9):
        ws_recap.cell(r, col).font = Font(bold=True)
        ws_recap.cell(r, col).fill = PatternFill("solid", fgColor="FFDDDDDD")
        ws_recap.cell(r, col).border = _border()

    for i, w in enumerate([28, 20, 20, 14, 14, 10, 16, 16], 1):
        ws_recap.column_dimensions[get_column_letter(i)].width = w

    # ── 2. Une feuille par couleur (groupe), triée : couleur → façade → sous-type ──
    for g in data["groups"]:
        import re
        sheet_name = re.sub(r'[\\/*?\[\]:]', '_', g["name"])[:31]  # Excel forbidden chars
        ws = wb.create_sheet(sheet_name)
        ws.append([g["name"]])
        ws["A1"].font = Font(bold=True, size=12)

        headers = ["Façade", "Sous-type", "Largeur (mm)", "Hauteur (mm)",
                   "Quantité", "Surface unitaire (m²)", "Surface totale (m²)"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(2, col).font = Font(bold=True)
            ws.cell(2, col).fill = PatternFill("solid", fgColor="FFE8E8E8")
            ws.cell(2, col).border = _border()

        total_surface = 0.0
        fill = _header_fill(g["color"])

        for ss in g["subsections"]:
            for st in ss["panelSubtypes"]:
                for p in st["pieces"]:
                    c1, c2 = p["w"], p["h"]
                    s_unit = round(c1 * c2 / 1_000_000, 4)
                    s_tot = round(s_unit * p["qty"], 4)
                    total_surface += s_tot
                    ws.append([ss["name"], st["name"], c1, c2, p["qty"], s_unit, s_tot])
                    r = ws.max_row
                    for col in range(1, 8):
                        ws.cell(r, col).border = _border()
                    ws.cell(r, 1).fill = fill

        # Ligne total
        ws.append(["TOTAL " + g["name"], None, None, None, None, None, round(total_surface, 4)])
        r = ws.max_row
        for col in range(1, 8):
            ws.cell(r, col).font = Font(bold=True)
            ws.cell(r, col).fill = PatternFill("solid", fgColor="FFDDDDDD")
            ws.cell(r, col).border = _border()

        for i, w in enumerate([20, 20, 14, 14, 10, 16, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 3. Feuille Ossature ────────────────────────────────────────────────
    ossature = data.get("ossature_facades", {})
    if ossature:
        ws_oss = wb.create_sheet("Ossature")
        title = data["chantier"]["nom"] + " — Ossature"
        ws_oss.append([title])
        ws_oss["A1"].font = Font(bold=True, size=13)
        ws_oss.append([])  # ligne vide

        # En-têtes
        ws_oss.append(["Façade", "Oméga (ml)", "Zed (ml)", "Total (ml)"])
        r = ws_oss.max_row
        for col in range(1, 5):
            ws_oss.cell(r, col).font = Font(bold=True)
            ws_oss.cell(r, col).fill = PatternFill("solid", fgColor="FFE0E0E0")
            ws_oss.cell(r, col).border = _border()

        grand_omega = 0
        grand_zed = 0

        for facade_name in facades_order:
            oss = ossature.get(facade_name, {})
            omega_ml = oss.get("omega_ml", 0)
            zed_ml = oss.get("zed_ml", 0)
            total_ml = round(omega_ml + zed_ml, 2)
            grand_omega += omega_ml
            grand_zed += zed_ml

            ws_oss.append([facade_name, omega_ml, zed_ml, total_ml])
            r = ws_oss.max_row
            for col in range(1, 5):
                ws_oss.cell(r, col).border = _border()
            ws_oss.cell(r, 2).number_format = '0.00'
            ws_oss.cell(r, 3).number_format = '0.00'
            ws_oss.cell(r, 4).number_format = '0.00'

        # Ligne total
        ws_oss.append(["TOTAL", round(grand_omega, 2), round(grand_zed, 2),
                        round(grand_omega + grand_zed, 2)])
        r = ws_oss.max_row
        for col in range(1, 5):
            ws_oss.cell(r, col).font = Font(bold=True)
            ws_oss.cell(r, col).fill = PatternFill("solid", fgColor="FFDDDDDD")
            ws_oss.cell(r, col).border = _border()
            if col >= 2:
                ws_oss.cell(r, col).number_format = '0.00'

        # Section détaillée par hauteur de profil
        for profile_type, label, color_hex, fill_hex in [
            ("omega_details", "DÉTAIL OMÉGA PAR HAUTEUR", "FF2563EB", "FFD4E6FF"),
            ("zed_details", "DÉTAIL ZED PAR HAUTEUR", "FFCA8A04", "FFFFF3CD"),
        ]:
            ws_oss.append([])
            ws_oss.append([label])
            r = ws_oss.max_row
            ws_oss.cell(r, 1).font = Font(bold=True, size=11, color=color_hex)

            ws_oss.append(["Hauteur (mm)", "Quantité", "Total (ml)"])
            r = ws_oss.max_row
            for col in range(1, 4):
                ws_oss.cell(r, col).font = Font(bold=True)
                ws_oss.cell(r, col).fill = PatternFill("solid", fgColor=fill_hex)
                ws_oss.cell(r, col).border = _border()

            # Agréger toutes les façades
            all_details = defaultdict(int)
            for facade_name in facades_order:
                oss = ossature.get(facade_name, {})
                for h_str, qty in oss.get(profile_type, {}).items():
                    all_details[int(h_str)] += qty

            for h in sorted(all_details.keys(), reverse=True):
                qty = all_details[h]
                ml = round(h * qty / 1000, 2)
                ws_oss.append([h, qty, ml])
                r = ws_oss.max_row
                for col in range(1, 4):
                    ws_oss.cell(r, col).border = _border()

        ws_oss.column_dimensions["A"].width = 22
        ws_oss.column_dimensions["B"].width = 16
        ws_oss.column_dimensions["C"].width = 16
        ws_oss.column_dimensions["D"].width = 16

    wb.save(out_path)
    return out_path

# ─── Arrondissement d'un JSON existant ───────────────────────────────────────

def round_json(data):
    """Arrondit toutes les dimensions w/h au mm le plus proche dans un JSON existant."""
    import copy
    d = copy.deepcopy(data)
    for g in d.get("groups", []):
        for ss in g.get("subsections", []):
            for st in ss.get("panelSubtypes", []):
                for p in st.get("pieces", []):
                    p["w"] = round_mm(p["w"])
                    p["h"] = round_mm(p["h"])
    return d

# ─── CLI ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Parse DXF/DWG → JSON + Excel Calepinage")
    parser.add_argument("input", help="Fichier DXF ou DWG en entrée")
    parser.add_argument("--excel", "-e", help="Chemin de sortie Excel (.xlsx)", default=None)
    parser.add_argument("--out", "-o", help="Chemin de sortie JSON", default=None)
    parser.add_argument("--round-json", help="Arrondir un JSON existant (ne parse pas de DXF)")
    args = parser.parse_args()

    # Mode arrondi d'un JSON existant
    if args.round_json:
        with open(args.round_json, encoding="utf-8") as f:
            data = json.load(f)
        data = round_json(data)
        out = args.out or args.round_json
        with open(out, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"JSON arrondi sauvegardé → {out}", file=sys.stderr)
        return

    filepath = args.input
    tmp_dxf = None

    # Si DWG : conversion
    if filepath.lower().endswith(".dwg"):
        print("Conversion DWG → DXF...", file=sys.stderr)
        try:
            tmp_dxf = dwg_to_dxf(filepath)
            filepath = tmp_dxf
        except RuntimeError as e:
            print(f"ERREUR conversion DWG : {e}", file=sys.stderr)
            sys.exit(1)

    # Parse DXF
    print(f"Lecture de {filepath}...", file=sys.stderr)
    data = parse_dxf_file(filepath)

    # JSON
    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(json_str)
        print(f"JSON sauvegardé → {args.out}", file=sys.stderr)
    else:
        print(json_str)

    # Excel
    excel_path = args.excel
    if excel_path is None and args.out:
        excel_path = os.path.splitext(args.out)[0] + ".xlsx"
    if excel_path:
        generate_excel(data, excel_path)
        print(f"Excel sauvegardé → {excel_path}", file=sys.stderr)

    # Nettoyage DXF temporaire
    if tmp_dxf and os.path.exists(tmp_dxf):
        os.unlink(tmp_dxf)

if __name__ == "__main__":
    main()
